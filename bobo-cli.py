import pandas as pd
import argparse
import os
import shutil
from datetime import datetime
from openpyxl import load_workbook

def load_aisle_data(excel_path):
    xls = pd.ExcelFile(excel_path)
    aisle_data = []
    sheet_data = {}

    for sheet_name in xls.sheet_names:
        if sheet_name.lower() == "aisle 1":
        # if sheet_name.lower().startswith("aisle"):
            print(f"Loading data from sheet: {sheet_name}")
            df = pd.read_excel(xls, sheet_name=sheet_name, header=1)
            # df = df.iloc[1:]  # Skip the header row
            try:
                # Identify all sections with 'Mashgin ID #' and 'Sales $' columns
                for col_index, col_name in enumerate(df.columns):
                    if "Mashgin ID #" in col_name:
                        sales_col_index = col_index + 1  # Assume 'Sales $' is the next column
                        if sales_col_index < len(df.columns) and "Sales $" in df.columns[sales_col_index]:
                            section = df[[col_name, df.columns[sales_col_index]]].rename(
                                columns={col_name: "Mashgin ID", df.columns[sales_col_index]: "Sales $"}
                            )
                            aisle_data.append(section)
                sheet_data[sheet_name] = df  # Store the original sheet data for later updates
            except Exception as e:
                print(f"Warning: Skipped sheet '{sheet_name}' due to error: {e}")

    all_aisle_data = pd.concat(aisle_data, ignore_index=True).dropna().reset_index(drop=True)
    all_aisle_data["Mashgin ID"] = all_aisle_data["Mashgin ID"].astype(str)
    return all_aisle_data, sheet_data

def load_import_csv(import_path):
    print(f"Loading import CSV from: {import_path}")
    
    # Read the CSV file, skipping rows until "ITEM SALES" is found
    with open(import_path, "r") as file:
        lines = file.readlines()
    
    # Find the starting line for the actual data
    start_index = 0
    for i, line in enumerate(lines):
        if "ITEM SALES" in line:
            start_index = i + 1  # Data starts after "ITEM SALES"
            break
    
    # Load the data into a DataFrame, skipping the metadata rows
    import_df = pd.read_csv(
        import_path,
        skiprows=start_index,
        thousands=',',
        dtype={'Pos Id': 'str', 'Count': 'int'},
        converters={'Total': convert_currency_to_float},
        encoding="utf-8"
    )

    # Strip whitespace from column names and rename for consistency
    import_df.columns = [col.strip() for col in import_df.columns]

    import_df = import_df.rename(columns={"Pos Id": "Mashgin ID", "Total": "Sales $"})
    import_df["Mashgin ID"] = import_df["Mashgin ID"].apply(lambda x: str(int(x)).strip() if pd.notnull(x) and x != "" else "")

    print(f"Loaded {len(import_df)} rows from import CSV.")

    # load first 5 rows for debugging
    print(import_df.head())
    # Check for duplicate Mashgin IDs
    if import_df["Mashgin ID"].duplicated().any():
        print("Warning: Duplicate Mashgin IDs found in import CSV.")
        duplicate_count = import_df["Mashgin ID"].duplicated(keep=False).sum()
        if duplicate_count > 0:
            print(f"Found {duplicate_count} duplicate Mashgin IDs. Combining duplicates.")
        
        import_df = import_df.groupby("Mashgin ID", as_index=False).agg({
            "Sales $": "sum",
            "Count": "sum"
        })
    print(f"Processed {len(import_df)} rows after combining duplicates.")
    
    return import_df

def merge_data(aisle_df, import_df):
    # Merge and add "Total Sales $" column
    merged = aisle_df.copy()

    for col_index, col_name in enumerate(aisle_df.columns):
        if "Mashgin ID" in col_name:
            sales_col_index = col_index + 1  # Assume 'Sales $' is the next column
            if sales_col_index < len(aisle_df.columns) and "Sales $" in aisle_df.columns[sales_col_index]:
                print(f"Processing Mashgin ID column: {col_name} and Sales $ column: {aisle_df.columns[sales_col_index]}")
                
                # Extract the relevant section
                section = aisle_df[[col_name, aisle_df.columns[sales_col_index]]].rename(
                    columns={col_name: "Mashgin ID", aisle_df.columns[sales_col_index]: "Sales $"}
                )
                
                # Merge with import_df to update Sales $
                updated_section = pd.merge(
                    section,
                    import_df[["Mashgin ID", "Sales $"]],
                    on="Mashgin ID",
                    how="left",
                    suffixes=("", "_updated")
                )
                
                # Update the Sales $ column with the merged values
                updated_section["Sales $"] = updated_section["Sales $_updated"].combine_first(updated_section["Sales $"])
                updated_section = updated_section.drop(columns=["Sales $_updated"])
                
                # Update the original DataFrame
                merged[aisle_df.columns[sales_col_index]] = updated_section["Sales $"].values

    print("All Mashgin ID columns and corresponding Sales $ columns have been updated.")

    # Show the first 5 rows of the merged DataFrame for debugging
    print(merged.head())

    if merged["Sales $"].isnull().any():
        print("Warning: Some Mashgin IDs did not match and have missing Sales $ values.")
    

    if merged["Mashgin ID"].duplicated().any():
        print("Warning: Duplicate Mashgin IDs found in merged data.")
    
    print(f"Successfully merged {len(merged)} rows.")
    
    return merged

def backup_file(filepath):
    backup_path = filepath + ".bak"
    shutil.copy2(filepath, backup_path)
    print(f"✅ Backup created: {backup_path}")

def convert_currency_to_float(value):
    """Converts a string representing currency to a float."""
    if isinstance(value, str):
        return float(value.replace('$', '').replace(',', ''))
    return float(value)

def save_report(merged_df, output_dir, excel_path, sheet_data):
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel_filename = f"updated_{os.path.basename(excel_path)}"
    output_excel_path = os.path.join(output_dir, output_excel_filename)
    
    # Copy the original Excel file to the output directory
    shutil.copy2(excel_path, output_excel_path)
    print(f"✅ Original Excel file copied to: {output_excel_path}")
    
    # Load the copied Excel file with openpyxl
    workbook = load_workbook(output_excel_path)
    
    for sheet_name, df in sheet_data.items():
        if sheet_name.lower().startswith("aisle"):
            sheet = workbook[sheet_name]
            
            # Unmerge all merged cells in the sheet
            merged_cells = list(sheet.merged_cells)
            for merged_cell in merged_cells:
                sheet.unmerge_cells(str(merged_cell))
            
            # Update each section with the "Total Sales $" column
            for col_index, col_name in enumerate(df.columns):
                if "Mashgin ID#" in col_name:
                    sales_col_index = col_index + 1  # Assume 'Sales $' is the next column
                    if sales_col_index < len(df.columns) and "Sales $" in df.columns[sales_col_index]:
                        # Extract the section with Mashgin ID and Sales $
                        section = df[[col_name, df.columns[sales_col_index]]].rename(
                            columns={col_name: "Mashgin ID", df.columns[sales_col_index]: "Sales $"}
                        )
                        # Merge the section with the merged data to get updated "Total Sales $"
                        updated_section = pd.merge(
                            section, 
                            merged_df[["Mashgin ID", "Total Sales $"]], 
                            on="Mashgin ID", 
                            how="left"
                        )
                        # Update the original DataFrame with the new "Total Sales $" values
                        df[df.columns[sales_col_index]] = updated_section["Total Sales $"].fillna("").values
            
            # Write the updated DataFrame back to the sheet
            for row_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start at row 2 to skip the header
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Re-merge cells if necessary (optional, based on your requirements)
            # Example: sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    
    # Save the updated workbook
    workbook.save(output_excel_path)
    print(f"✅ Updated Excel file saved to: {output_excel_path}")

    # Save the merged data as a separate report
    report_filename = f"report_{timestamp}.xlsx"
    report_filepath = os.path.join(output_dir, report_filename)
    merged_df.to_excel(report_filepath, index=False, engine="openpyxl")
    print(f"✅ Merged report saved to: {report_filepath}")

def main():
    parser = argparse.ArgumentParser(
        prog="bobo-cli",
        description="🛒 Merge Aisle Excel sheets with Import CSV by Mashgin ID (Windows-friendly)."
    )

    parser.add_argument("-e", "--excel_file", required=True, help="Path to Excel file (with Aisle sheets)")
    parser.add_argument("-i", "--import_csv", required=True, help="Path to Import CSV file (with Mashgin IDs and Total)")
    parser.add_argument("-o", "--output_dir", required=True, help="Output folder for report")
    args = parser.parse_args()


    if not os.path.exists(args.excel_file):
        print(f"❌ Excel file not found: {args.excel_file}")
        return
    if not os.path.exists(args.import_csv):
        print(f"❌ Import CSV not found: {args.import_csv}")
        return

    try:
        backup_file(args.excel_file)
        aisle_data, sheet_data = load_aisle_data(args.excel_file)
        import_data = load_import_csv(args.import_csv)
        merged = merge_data(aisle_data, import_data)
        save_report(merged, args.output_dir, args.excel_file, sheet_data)
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()